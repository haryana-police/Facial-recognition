import sys
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

# Load Excel dd_nos
wb = openpyxl.load_workbook('D:/UIDB_CCTNS.xlsx', read_only=True, data_only=True)
ws = wb.active
headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(max_row=1))]
dd_col = headers.index('dd_no') if 'dd_no' in headers else 0
excel_ddnos = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    v = str(row[dd_col]).strip() if row[dd_col] else ''
    if v and v != 'None': excel_ddnos.add(v)
wb.close()
print(f'Excel dd_nos count: {len(excel_ddnos)}')

# Load image folder names
img_dir = Path('D:/images')
img_folders = {f.name for f in img_dir.iterdir() if f.is_dir()}
print(f'Image folders count: {len(img_folders)}')

# Check matches
matches = excel_ddnos.intersection(img_folders)
print(f'Matching dd_nos: {len(matches)}')
print()
print('Sample image folders:', list(img_folders)[:5])
print()
print('Sample excel dd_nos:', list(excel_ddnos)[:5])
