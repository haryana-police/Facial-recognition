# -*- coding: utf-8 -*-
"""
import_from_excel.py
──────────────────────────────────────────────────────────────────
UIDB_CCTNS.xlsx aur D:\\images folder se suspect data import karta hai.

Image matching logic:
  D:\\images\\<dd_no>\\face_front.jpg  →  matched by folder name = dd_no

Usage:
  python scripts/import_from_excel.py

  Ya custom paths ke saath:
  python scripts/import_from_excel.py "D:\\images" "D:\\UIDB_CCTNS.xlsx"

Features:
  - Resume safe: already imported dd_no ko skip karta hai
  - Excel (.xlsx) directly padh sakta hai
  - Hindi/Unicode text support
  - Har 50 records pe progress save hota hai
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')
import sqlite3
import shutil
import requests
import openpyxl
from pathlib import Path
from datetime import datetime

# ── Default Config ─────────────────────────────────────────────────
IMAGES_DIR    = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('D:/images')
EXCEL_PATH    = Path(sys.argv[2]) if len(sys.argv) > 2 else Path('D:/UIDB_CCTNS.xlsx')
DB_PATH       = Path('forensic_suspects.db')
PHOTOS_DIR    = Path('suspect_photos')
PYTHON_AI_URL = 'http://127.0.0.1:8000/api/v1/forensic-extract'
NODE_RELOAD   = 'http://127.0.0.1:8080/api/reload-db'
IMAGE_NAME    = 'face_front.jpg'   # image filename inside each dd_no folder
# ───────────────────────────────────────────────────────────────────

PHOTOS_DIR.mkdir(exist_ok=True)


def safe(val):
    """Convert any value to string, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() != 'none' else None


def safe_int(val):
    try:
        v = int(float(str(val)))
        return v if v > 0 else 0
    except:
        return 0


def safe_float(val):
    try:
        v = float(str(val))
        return v if v > 0 else 0.0
    except:
        return 0.0


def load_excel(path):
    """Load Excel file and return (headers, rows_dict_by_dd_no)."""
    print(f'[EXCEL] Loading {path} ...')
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    raw_headers = [str(cell.value).strip().lower().replace(' ', '_')
                   if cell.value else '' for cell in next(ws.iter_rows(max_row=1))]
    print(f'[EXCEL] Columns: {raw_headers}')

    rows = {}
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Map header -> value
        record = {raw_headers[i]: row[i] for i in range(len(raw_headers)) if i < len(row)}
        dd = safe(record.get('dd_no', ''))
        if dd:
            rows[dd] = record
            total += 1

    wb.close()
    print(f'[EXCEL] Loaded {total} rows with dd_no\n')
    return rows


def extract_embedding(img_path):
    """Call Python AI server to get 512-D ArcFace embedding."""
    with open(str(img_path), 'rb') as f:
        resp = requests.post(
            PYTHON_AI_URL,
            files={'file': (img_path.name, f, 'image/jpeg')},
            data={'fidelity_w': '0.85'},
            timeout=120
        )
    if resp.status_code != 200:
        raise Exception(f'AI Error {resp.status_code}: {resp.text[:200]}')
    return resp.json()['embedding']


def main():
    # ── Validate paths ─────────────────────────────────────────────
    if not IMAGES_DIR.is_dir():
        print(f'ERROR: Images folder not found: {IMAGES_DIR}')
        sys.exit(1)
    if not EXCEL_PATH.exists():
        print(f'ERROR: Excel file not found: {EXCEL_PATH}')
        sys.exit(1)
    if not DB_PATH.exists():
        print(f'ERROR: Database not found: {DB_PATH}')
        sys.exit(1)

    # ── Load Excel ─────────────────────────────────────────────────
    excel_data = load_excel(EXCEL_PATH)

    # ── Scan image folders ─────────────────────────────────────────
    # Each subfolder name = dd_no, contains face_front.jpg
    image_folders = [f for f in IMAGES_DIR.iterdir() if f.is_dir()]
    print(f'[SCAN] Found {len(image_folders)} image folders in {IMAGES_DIR}')

    conn = sqlite3.connect(str(DB_PATH))

    success_count = 0
    skip_count    = 0
    no_excel_match = 0
    no_image      = 0
    fail_count    = 0

    print(f'[START] Processing...\n{"="*60}')

    for folder in image_folders:
        dd_no     = folder.name  # folder name IS the dd_no
        img_path  = folder / IMAGE_NAME

        # Check image exists
        if not img_path.exists():
            # Try any image in the folder
            imgs = list(folder.glob('*.jpg')) + list(folder.glob('*.jpeg')) + list(folder.glob('*.png'))
            if not imgs:
                print(f'[SKIP] No image in folder: {dd_no}')
                no_image += 1
                continue
            img_path = imgs[0]

        # Skip if already in DB
        existing = conn.execute('SELECT id FROM suspect WHERE dd_no = ?', (dd_no,)).fetchone()
        if existing:
            skip_count += 1
            continue

        # Get metadata from Excel
        meta = excel_data.get(dd_no, {})
        if not meta:
            print(f'[WARN] No Excel row for dd_no={dd_no}, will import with image only')
            no_excel_match += 1

        print(f'[PROCESS] dd_no={dd_no}  |  image={img_path.name}')

        # Get next DB ID
        max_id  = conn.execute('SELECT COALESCE(MAX(id), 0) FROM suspect').fetchone()[0]
        next_id = max_id + 1

        # Copy image to suspect_photos/
        dd_safe      = dd_no.replace('/', '_').replace('\\', '_')[:50]
        new_filename = f'{next_id}_{dd_safe}.jpg'
        dest_path    = PHOTOS_DIR / new_filename
        try:
            shutil.copy2(str(img_path), str(dest_path))
        except Exception as e:
            print(f'  [ERROR] Copy failed: {e}')
            fail_count += 1
            continue

        relative_path = f'suspect_photos/{new_filename}'

        # Extract embedding
        try:
            emb_list = extract_embedding(img_path)
            emb_csv  = ','.join(f'{v:.6f}' for v in emb_list)
        except requests.exceptions.ConnectionError:
            print('\n[ERROR] Python AI server not running on port 8000!')
            print('        Start it with: python main.py')
            dest_path.unlink(missing_ok=True)
            conn.close()
            sys.exit(1)
        except Exception as e:
            print(f'  [ERROR] Embedding failed: {e}')
            dest_path.unlink(missing_ok=True)
            fail_count += 1
            continue

        # Insert into DB
        try:
            conn.execute(
                '''INSERT INTO suspect (
                    name, image_path, embedding_vector,
                    dd_no, found_date, found_district, ps_name, found_loc,
                    gender, age_min, age_max, height_cm, build,
                    skin_tone, hair_color, beard, visible_marks,
                    clothing_description, notes, additional_details
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                [
                    safe(meta.get('name')) or 'Unknown',
                    relative_path,
                    emb_csv,
                    dd_no,
                    safe(meta.get('found_date')),
                    safe(meta.get('found_district')),
                    safe(meta.get('ps_name')),
                    safe(meta.get('found_loc')),
                    safe(meta.get('gender')),
                    safe_int(meta.get('age_min')),
                    safe_int(meta.get('age_max')),
                    safe_float(meta.get('height_cm')),
                    safe(meta.get('build')),
                    safe(meta.get('skin_tone')),
                    safe(meta.get('hair_color')),
                    safe(meta.get('beard')),
                    safe(meta.get('visible_marks')),
                    safe(meta.get('clothing_description')),
                    safe(meta.get('notes')),
                    safe(meta.get('additional_details')),
                ]
            )
            conn.commit()
            success_count += 1
            print(f'  [OK] Saved id={next_id}')

        except Exception as e:
            print(f'  [ERROR] DB insert failed: {e}')
            dest_path.unlink(missing_ok=True)
            fail_count += 1

    conn.close()

    # ── Summary ────────────────────────────────────────────────────
    print(f'\n' + '='*60)
    print(f'  IMPORT COMPLETE')
    print('='*60)
    print(f'  Total Folders  : {len(image_folders)}')
    print(f'  Successful     : {success_count}')
    print(f'  Skipped (dup)  : {skip_count}')
    print(f'  No Excel match : {no_excel_match}')
    print(f'  No image found : {no_image}')
    print(f'  Failed         : {fail_count}')
    print('='*60)

    # Reload Node.js DB
    if success_count > 0:
        try:
            r = requests.post(NODE_RELOAD, timeout=5)
            cnt = r.json().get('suspectCount', '?')
            print(f'\n[OK] Node.js DB reloaded! Total suspects: {cnt}')
        except Exception:
            print('\n[NOTE] Could not reload Node.js DB. Restart node_backend if needed.')


if __name__ == '__main__':
    main()
